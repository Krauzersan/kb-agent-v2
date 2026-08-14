"""Excel-отчёт по метрикам и аналитике — для админки (см. admin.py: /admin/export.xlsx).

Листы: обзор (с мини-сводкой за последние 7 дней — по сути еженедельный дайджест,
только скачиваете вручную, когда удобно, а не по расписанию), «За неделю» (лог за
последние 7 дней), «Проблемные ответы» (низкая оценка или ответ без опоры на базу —
то же определение «пробела», что и на странице «Аналитика»), и полный лог вопросов.
Каждый лист — настоящая Excel-таблица (фильтры, полосатая заливка, закреплённая
шапка), чтобы отчётом было удобно пользоваться сразу после скачивания, без
дополнительного форматирования руками.
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

_LOW_RATING_MAX = 5  # 1-10; всё что <= этого — «низкая оценка» для отчёта
_WEEK_DAYS = 7

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")

_RATING_FILLS = {
    "low": PatternFill("solid", fgColor="FCA5A5"),      # 1-5 — красный
    "mid": PatternFill("solid", fgColor="FDE68A"),       # 6-7 — жёлтый
    "high": PatternFill("solid", fgColor="86EFAC"),      # 8-10 — зелёный
    "none": PatternFill("solid", fgColor="E5E7EB"),      # не оценено — серый
}


def _rating_fill(rating) -> PatternFill:
    if rating is None:
        return _RATING_FILLS["none"]
    if rating <= 5:
        return _RATING_FILLS["low"]
    if rating <= 7:
        return _RATING_FILLS["mid"]
    return _RATING_FILLS["high"]


def _problem_reason(row: dict) -> str:
    reasons = []
    if not row.get("sources"):
        reasons.append("нет источников в базе")
    rating = row.get("rating")
    if rating is not None and rating <= _LOW_RATING_MAX:
        reasons.append(f"низкая оценка ({rating})")
    return "; ".join(reasons)


def _is_problem(row: dict) -> bool:
    rating = row.get("rating")
    return (not row.get("sources")) or (rating is not None and rating <= _LOW_RATING_MAX)


def _week_rows(rows: list) -> list:
    """Строки за последние 7 дней. created_at — "YYYY-MM-DD HH:MM:SS" (UTC, db._now()),
    формат сортируется как строка, поэтому парсить дату не нужно."""
    cutoff = (datetime.now(timezone.utc) - timedelta(days=_WEEK_DAYS)).strftime("%Y-%m-%d %H:%M:%S")
    return [r for r in rows if (r.get("created_at") or "") >= cutoff]


def _mini_stats(rows: list) -> dict:
    rated = [r["rating"] for r in rows if r.get("rating") is not None]
    return {
        "questions": len(rows),
        "rated": len(rated),
        "avg_rating": round(sum(rated) / len(rated), 1) if rated else None,
        "problems": sum(1 for r in rows if _is_problem(r)),
    }


_LOG_COLUMNS = [
    ("ID", 6), ("Дата", 17), ("Канал", 10), ("Тема", 22),
    ("Кто спросил", 18), ("Вопрос", 45), ("Ответ", 55),
    ("Оценка", 9), ("Источников", 11), ("Проблема", 28),
]


def _write_log_sheet(ws, rows: list, title: str) -> None:
    ws.title = title
    for col, (name, width) in enumerate(_LOG_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    for i, row in enumerate(rows, start=2):
        rating = row.get("rating")
        values = [
            row.get("id"),
            (row.get("created_at") or "")[:19].replace("T", " "),
            row.get("channel") or "",
            row.get("topic") or "",
            row.get("asker_name") or "",
            row.get("question") or "",
            row.get("answer") or "",
            rating if rating is not None else "не оценено",
            len(row.get("sources") or []),
            _problem_reason(row) or "—",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.alignment = _WRAP if col in (6, 7) else _TOP
            if col == 8:
                cell.fill = _rating_fill(rating)

    last_row = max(2, len(rows) + 1)
    last_col = get_column_letter(len(_LOG_COLUMNS))
    if rows:
        table = Table(displayName=f"tbl_{title.replace(' ', '_')[:20]}",
                      ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
        )
        ws.add_table(table)


def _write_overview_sheet(ws, overview: dict, topic_stats: list, user_stats: list,
                          problem_count: int, total: int, week_stats: dict,
                          generated_at: str) -> None:
    ws.title = "Обзор"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    def header(text, row):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=13)
        return row + 1

    r = 1
    ws.cell(row=r, column=1, value="Отчёт по метрикам и аналитике").font = Font(bold=True, size=15)
    r += 1
    ws.cell(row=r, column=1, value=f"Сформирован: {generated_at}").font = Font(italic=True, color="6B7280")
    r += 2

    r = header(f"За последние {_WEEK_DAYS} дней", r)
    week_avg = week_stats.get("avg_rating")
    for label, value in [
        ("Вопросов агенту", week_stats.get("questions") or 0),
        ("Из них с оценкой (1-10)", week_stats.get("rated") or 0),
        ("Средняя оценка", week_avg if week_avg is not None else "—"),
        ("Проблемных ответов", week_stats.get("problems") or 0),
    ]:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=value)
        cell.alignment = Alignment(horizontal="right")
        r += 1
    r += 1

    r = header("За всё время", r)
    rated = overview.get("rated") or 0
    avg = overview.get("avg_rating")
    stats_rows = [
        ("Всего вопросов агенту", total),
        ("Из них с оценкой (1-10)", rated),
        ("Средняя оценка", avg if avg is not None else "—"),
        ("Проблемных ответов (низкая оценка или нет источников)", problem_count),
    ]
    for label, value in stats_rows:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=value)
        cell.alignment = Alignment(horizontal="right")
        r += 1
    r += 1

    if user_stats:
        r = header("По сотрудникам (кто спрашивает и как оценивает)", r)
        for col, name in enumerate(["Сотрудник", "Вопросов", "С оценкой", "Средняя оценка"], start=1):
            cell = ws.cell(row=r, column=col, value=name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        r += 1
        for u in user_stats:
            ws.cell(row=r, column=1, value=u.get("asker_name") or "Неизвестно")
            ws.cell(row=r, column=2, value=u.get("questions") or 0)
            ws.cell(row=r, column=3, value=u.get("rated") or 0)
            ws.cell(row=r, column=4, value=u.get("avg_rating") if u.get("avg_rating") is not None else "—")
            r += 1
        r += 1

    if topic_stats:
        r = header("По темам (о чём чаще всего спрашивают)", r)
        for col, name in enumerate(
            ["Тема", "Вопросов", "С оценкой", "Средняя оценка", "Без источников"], start=1
        ):
            cell = ws.cell(row=r, column=col, value=name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        r += 1
        for t in topic_stats:
            ws.cell(row=r, column=1, value=t.get("topic") or "")
            ws.cell(row=r, column=2, value=t.get("questions") or 0)
            ws.cell(row=r, column=3, value=t.get("rated") or 0)
            ws.cell(row=r, column=4, value=t.get("avg_rating") if t.get("avg_rating") is not None else "—")
            no_src = t.get("no_sources") or 0
            cell = ws.cell(row=r, column=5, value=no_src)
            if no_src:
                cell.fill = _RATING_FILLS["low"]
            r += 1


def build_report(overview: dict, user_stats: list, topic_stats: list, rows: list) -> bytes:
    """Собирает .xlsx в память и возвращает его байты — без записи на диск."""
    problems = [r for r in rows if _is_problem(r)]
    week = _week_rows(rows)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()
    _write_overview_sheet(wb.active, overview, topic_stats, user_stats,
                          problem_count=len(problems), total=len(rows),
                          week_stats=_mini_stats(week), generated_at=generated_at)
    _write_log_sheet(wb.create_sheet(), week, f"За {_WEEK_DAYS} дней")
    _write_log_sheet(wb.create_sheet(), problems, "Проблемные ответы")
    _write_log_sheet(wb.create_sheet(), rows, "Все вопросы")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
